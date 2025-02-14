import flet as ft
import logging
from typing import Optional, Callable, List
import locale
import asyncio
import openpyxl
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

from acessorios import BancoDeDados, Dialogo, BuscarCep, JanelaNotificacao
from pagina_dash import PaginaDashboard
from modelos import ModeloFornecedor, ModeloItem
from controles import ControleItem, ControleFornecedor, ControleGradeItem, ControlePagina, ControleVisualizacao
from pagina_lista_compras import PaginaListacompras
from pagina_fornecedores import PaginaFornecedores
from pagina_itens import PaginaItens

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)
# logging.disable(logging.CRITICAL)

locale.setlocale(locale.LC_ALL, "pt_BR.UTF-8")


class BarraPesquisa(ft.Container):
    def __init__(self, controle_grade: ControleGradeItem) -> None:
        super().__init__()
        self.controle_grade = controle_grade
        self.field_item = ft.TextField(
            width=250,
            label="Nome do item",
            border="underline",
            on_submit=self.filtrar
        )
        self.content = ft.Row([
            self.field_item,
            ft.IconButton(icon=ft.Icons.SEARCH, on_click=self.filtrar, icon_color=ft.Colors.BLACK87)
        ])

    def filtrar(self, e: ft.ControlEvent) -> None:
        self.controle_grade.filtrar(self.field_item.value)
        self.limpar_field()

    def limpar_field(self) -> None:
        self.field_item.value = None
        self.field_item.update()


class JanelaAdcionarItem(ft.AlertDialog):
    def __init__(self, controle_grade: ControleGradeItem) -> None:
        super().__init__(
            modal=True,
            title=ft.Text(value="Adicionar Item")
        )
        self.dialogo = Dialogo()
        self.controle_grade = controle_grade
        self.criar_conteudo()

    def criar_conteudo(self) -> None:
        self.criar_entradas()
        self.criar_botoes()
        self.content = ft.Stack(
            [
                ft.Column([
                    ft.Row([
                        self.entradas[0],
                        self.entradas[2]
                    ]),
                    ft.Row([
                        self.entradas[1],
                    ])
                ], spacing=10),
                self.dialogo
            ],
            width=450,
            height=120,
            alignment=ft.alignment.center
        )
        self.actions = self.botoes

    def criar_entradas(self) -> None:
        lista_categorias = [
            "Carnes", "Condimentos", "Frios", "Verdura", "Legume", "Fruta", "Limpeza"
        ]

        self.entradas = [
            ft.TextField(label="Nome", width=265, border="underline"),
            self.criar_dropdown(
                "Medida",
                ["Quilograma", "Litro", "Unidade"],
                width=150
            ),
            self.criar_dropdown("Categoria", lista_categorias, width=170)
        ]

        for entrada in self.entradas:
            entrada.on_focus = self.alterar_para_cor_padrao

    def criar_botoes(self) -> None:
        self.botoes = [
            ft.TextButton(
                text="Salvar",
                on_click=self.salvar_item,
                style=ft.ButtonStyle(bgcolor={ft.ControlState.HOVERED: ft.Colors.GREEN_100})
            ),
            ft.TextButton(
                text="Cancelar",
                on_click=self.cancelar,
                style=ft.ButtonStyle(bgcolor={ft.ControlState.HOVERED: ft.Colors.RED_100})
            )
        ]

    def criar_dropdown(
        self,
        label: str,
        options: Optional[List[str]],
        width: int=150,
        on_change: Optional[Callable[[ft.ControlEvent], None]] = None,
        visible: bool = True
    ) -> ft.Dropdown:
        return ft.Dropdown(
            options=[
                ft.dropdown.Option(op)
                for op in options
            ],
            label=label,
            width=width,
            border="underline",
            visible=visible,
            on_change=on_change
        )

    async def salvar_item(self, e: ft.ControlEvent) -> None:
        if all([entrada.value for entrada in self.entradas if entrada.visible]): 
            await self.salvar_bd([entrada.value.lower() for entrada in self.entradas])
        else:
            await self.mostrar_erro_campos_vazios()

    async def salvar_bd(self, args: List[str]) -> None:
        controle = ControleItem(
            ModeloItem(nome=args[0], medida=args[1], categoria=args[2]),
            ControleVisualizacao([self.controle_grade, self.dialogo])
        )
        await controle.criar_item()
        self.limpar_campos()

    def mostrar_erro_campos_vazios(self) -> None:
        for entrada in self.entradas:
            if not entrada.value:
                entrada.label_style = ft.TextStyle(color=ft.Colors.RED)
                entrada.update()

    def limpar_campos(self) -> None:
        for campo in self.entradas:
            campo.value = None
            campo.update()

    def alterar_para_cor_padrao(self, e: ft.ControlEvent) -> None:
        e.control.label_style = ft.TextStyle(color=None)
        e.control.update()

    def cancelar(self, e: ft.ControlEvent) -> None:
        self.page.close(self)


class JanelaAdcionarFornecedor(ft.AlertDialog):
    def __init__(self) -> None:
        super().__init__(
            modal=True,
            title=ft.Text(value="Adicionar Fornecedor")
        )
        self.dialogo = Dialogo()
        self.configurar_layout()

    def configurar_layout(self) -> None:
        self.criar_entradas()
        self.criar_botoes()
        self.content = ft.Stack(
            [
                ft.Column([
                    ft.Row([
                        self.entradas[0],
                        self.entradas[1]
                    ]),
                    ft.Row([self.entradas[2]]),
                    ft.Divider(),
                    ft.Row([
                        self.entradas[3],
                        self.entradas[4]
                    ]),
                    ft.Row([
                        self.entradas[5],
                        self.entradas[6],
                        ft.IconButton(icon=ft.Icons.SEARCH, on_click=self.buscar_cep)
                    ]),
                    ft.Row([
                        self.entradas[7],
                        self.entradas[8]
                    ])
                ], spacing=10),
                self.dialogo
            ],
            width=450,
            height=330,
            alignment=ft.alignment.center
        )
        self.actions = self.botoes

    def criar_entradas(self) -> None:
        self.entradas = [
            ft.TextField(label="Nome", width=265, border="underline"),
            ft.TextField(label="Telefone", width=170, border="underline"),
            ft.TextField(label="Responsavel", width=300, border="underline"),
            ft.TextField(label="Logradouro", width=310, border="underline"),
            ft.TextField(label="Numero", width=125, border="underline"),
            ft.TextField(label="Bairro", width=255, border="underline"),
            ft.TextField(label="Cep", width=135, border="underline"),
            ft.TextField(label="Cidade", width=300, border="underline"),
            ft.TextField(label="Estado", width=135, border="underline")
        ]

        for entrada in self.entradas:
            entrada.on_focus = self.alterar_para_cor_padrao

    def criar_botoes(self) -> None:
        self.botoes = [
            ft.TextButton(
                text="Salvar",
                on_click=self.salvar_fornecedor,
                style=ft.ButtonStyle(bgcolor={ft.ControlState.HOVERED: ft.Colors.GREEN_100})
            ),
            ft.TextButton(
                text="Cancelar",
                on_click=self.cancelar,
                style=ft.ButtonStyle(bgcolor={ft.ControlState.HOVERED: ft.Colors.RED_100})
            )
        ]

    async def salvar_fornecedor(self, e: ft.ControlEvent) -> None:
        if all([entrada.value for entrada in self.entradas[:2] if entrada.visible]): 
            await self.salvar_bd([entrada.value.lower() for entrada in self.entradas])
        else:
            self.mostrar_erro_campos_vazios()

    async def salvar_bd(self, args: List[str]) -> None:
        args.insert(0, None)
        controle = ControleFornecedor(ModeloFornecedor(*args), self.dialogo)
        self.dialogo.salvando()
        await controle.adicionar_fornecedor()
        self.limpar_campos()

    def mostrar_erro_campos_vazios(self) -> None:
        for entrada in self.entradas:
            if not entrada.value:
                entrada.label_style = ft.TextStyle(color=ft.Colors.RED)
                entrada.update()

    def limpar_campos(self) -> None:
        for campo in self.entradas:
            campo.value = None
            campo.update()

    def alterar_para_cor_padrao(self, e: ft.ControlEvent) -> None:
        e.control.label_style = ft.TextStyle(color=None)
        e.control.update()

    def cancelar(self, e: ft.ControlEvent) -> None:
        self.page.close(self)

    def preencher_dados_endereco(self, dados: list) -> None:
        for i, l in zip([3, 5, 7, 8], ["street", "neighborhood", "city", "state"]):
            self.entradas[i].value = dados[l]
            self.entradas[i].update()

    async def buscar_cep(self, e: ft.ControlEvent) -> None:
        if self.entradas[6].value:
            buscador = BuscarCep()
            self.dialogo.generico(ft.Icons.SEARCH, "Procurando")
            dados = buscador.obter_dados(self.entradas[6].value)
            if dados:
                self.dialogo.generico(ft.Icons.CHECK, "Sucesso")
                self.preencher_dados_endereco(dados)
            else:
                self.dialogo.generico(ft.Icons.ERROR_OUTLINE, "Houve um erro")
            await asyncio.sleep(2)
            self.dialogo.limpar()
        else:
            self.dialogo.generico(ft.Icons.ERROR_OUTLINE, "Digite o Cep")
            await asyncio.sleep(2)
            self.dialogo.limpar()


class JanelaListaCompras(ft.AlertDialog):
    def __init__(self, controle_pagina: ControlePagina) -> None:
        super().__init__(modal=True)
        self.controle_pagina = controle_pagina
        self.dropdown_categoria = ft.Dropdown(label="Categoria", width=200, on_change=self.filtrar_categoria)
        self.tabela_produtos = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("id"), visible=False),
                ft.DataColumn(ft.Text("Nome")),
                ft.DataColumn(ft.Text("Adicionar"))
            ],
            col=12
        )
        self.tabela_compra = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("id"), visible=False),
                ft.DataColumn(ft.Text("Nome")),
                ft.DataColumn(ft.Text("Remover"))
            ],
            col=12
        )
        self.lista_compras = []
        self.criar_conteudo()

    def criar_conteudo(self) -> None:
        self.content = ft.Container(
            ft.Column([
                self.dropdown_categoria,
                ft.Container(
                    ft.ResponsiveRow([
                        ft.Column([
                            ft.Card(
                                ft.Container(
                                    ft.Column([
                                        ft.ResponsiveRow([
                                            self.tabela_produtos
                                        ])
                                    ], scroll=ft.ScrollMode.ALWAYS),
                                    padding=ft.padding.only(top=10, bottom=10), expand=True
                                ),expand=True, elevation=5
                            )
                        ], col=6),
                        ft.Column([
                            ft.Card(
                                ft.Container(
                                    ft.Column([
                                        ft.ResponsiveRow([
                                            self.tabela_compra
                                        ])
                                    ], scroll=ft.ScrollMode.ALWAYS),
                                    padding=ft.padding.only(top=10, bottom=10), expand=True
                                ),
                                expand=True, elevation=5
                            )
                        ], col=6)
                    ]), expand=True
                )
            ]), width=650
        )
        self.actions = [
            ft.TextButton(
                text="Criar Lista",
                on_click=self.criar_lista,
                style=ft.ButtonStyle(bgcolor={ft.ControlState.HOVERED: ft.Colors.GREEN_100})
            ),
            ft.TextButton(
                text="Cancelar",
                on_click=lambda e: self.page.close(self),
                style=ft.ButtonStyle(bgcolor={ft.ControlState.HOVERED: ft.Colors.RED_100})
            )
        ]

    def adicionar_linha_tabela_itens(self, indice: int) -> None:
        dados = self.dados[indice]
        self.tabela_produtos.rows.append(
            ft.DataRow(
                cells=[
                    ft.DataCell(ft.Text(dados[0]), visible=False),
                    ft.DataCell(ft.Text(dados[1])),
                    ft.DataCell(
                        ft.IconButton(
                            ft.Icons.ADD,
                            on_click=lambda e, id=dados[0]: self.adicionar_item_lista_compra(id)
                        )
                    )
                ]
            )
        )
        self.tabela_produtos.update()

    def adicionar_item_lista_compra(self, id: int) -> None:
        if id not in self.lista_compras:
            self.lista_compras.append(id)
            for dado in self.dados:
                if dado[0] == id:
                    self.tabela_compra.rows.append(
                        ft.DataRow(
                            cells=[
                                ft.DataCell(ft.Text(dado[0]), visible=False),
                                ft.DataCell(ft.Text(dado[1])),
                                ft.DataCell(
                                    ft.IconButton(
                                        ft.Icons.DELETE, on_click=lambda e, id=dado[0]: self.apagar_item_tabela_compra(id)
                                    )
                                )
                            ]
                        )
                    )
                    self.tabela_compra.update()

    def apagar_registros_tabela_produtos(self) -> None:
        self.tabela_produtos.rows.clear()
        self.tabela_produtos.update()

    def apagar_item_tabela_compra(self, id: int) -> None:
        self.lista_compras.remove(id)
        for linha in self.tabela_compra.rows:
            if linha.cells[0].content.value == id:
                self.tabela_compra.rows.remove(linha)
        self.tabela_compra.update()

    def filtrar_categoria(self, e: ft.ControlEvent) -> None:
        self.adicionar_registro_tabela_itens(e.data)

    def adicionar_registro_tabela_itens(self, categoria: str) -> None:
        self.apagar_registros_tabela_produtos()
        for i, dado in enumerate(self.dados):
            if categoria == dado[2]:
                self.adicionar_linha_tabela_itens(i)

    def adicionar_categorias_dropdown(self) -> None:
        if self.dados:
            cats = list(set(dado[2] for dado in self.dados))
            self.dropdown_categoria.options = [
                ft.dropdown.Option(value)
                for value in cats
            ]
            self.dropdown_categoria.value = cats[0]
            self.dropdown_categoria.update()
            self.adicionar_registro_tabela_itens(cats[0])

    async def ler_dados(self) -> None:
        bd = BancoDeDados("db_app6.db")
        self.dados = await bd.fetch_all("SELECT id, nome, categoria FROM produto;")
        self.adicionar_categorias_dropdown()

    def criar_lista(self, e: ft.ControlEvent) -> None:
        if self.tabela_compra.rows:
            infos_produtos = [
                (row.cells[0].content.value, row.cells[1].content.value)
                for row in self.tabela_compra.rows
            ]
            pagina = PaginaListacompras(infos_produtos)
            self.page.close(self)
            self.controle_pagina.alterar_para_barra_voltar()
            self.controle_pagina.add_acao_barra(
                ft.Container(
                    ft.IconButton(ft.Icons.CREATE, on_click=pagina.criar_planilha),
                    padding=ft.padding.only(right=20)
                )
            )
            self.controle_pagina.atualizar_pagina(pagina)

    def did_mount(self) -> None:
        self.page.run_task(self.ler_dados)


class PlanilhaCotacao:
    def __init__(self, produtos: ft.DataRow, fornecedores: ft.DataRow, nome_arquivo: str = "cotacao.xlsx") -> None:
        self.produtos = produtos
        self.fornecedores = fornecedores
        self.nome_arquivo = nome_arquivo
        self.colunas = ["produto"]
        self.book = openpyxl.Workbook()

    def criar(self) -> None:
        produtos = self.extrair_produtos()
        fornecedores = self.extrair_fornecedores()
        self.colunas.extend(fornecedores)

        ws = self.book.active
        ws.append(self.colunas)

        for produto in produtos:
            produto.extend(['-' for _ in range(len(fornecedores))])
            ws.append(produto)

        dim_holder = DimensionHolder(worksheet=ws)
        dims = [20]
        dims.extend([len(fornecedor) for fornecedor in fornecedores])

        for col in range(ws.min_column, ws.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=dims[col-1])
        
        ws.column_dimensions = dim_holder

        self.adicionar_data_ao_nome()
        self.book.save(self.nome_arquivo)

    def extrair_produtos(self) -> None:
        return [
            [linha.cells[0].content.value]
            for linha in self.produtos
        ]

    def extrair_fornecedores(self) -> None:
        return [
            linha.cells[0].content.value
            for linha in self.fornecedores
        ]
    
    def adicionar_data_ao_nome(self):
        nome_list = self.nome_arquivo.split(".")
        nome_full = nome_list[0] + f"_{datetime.now().strftime("%d_%m_%Y")}." + nome_list[1]
        self.nome_arquivo = nome_full


class JanelaCotacao(ft.AlertDialog):
    def __init__(self) -> None:
        super().__init__(modal=True)
        self.dropdown_categoria = ft.Dropdown(label="Categoria", width=200, on_change=self.filtrar_categoria)
        self.tabela_produtos = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("id"), visible=False),
                ft.DataColumn(ft.Text("Nome")),
                ft.DataColumn(ft.Text("Adicionar"))
            ],
            col=12
        )
        self.tabela_cotacao = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("id"), visible=False),
                ft.DataColumn(ft.Text("Nome")),
                ft.DataColumn(ft.Text("Remover"))
            ],
            col=12
        )
        self.dialogo = Dialogo()
        self.lista_cotacao = []
        self.criar_conteudo()

    def criar_conteudo(self) -> None:
        self.content = ft.Container(
            ft.Column([
                self.dropdown_categoria,
                ft.Container(
                    ft.ResponsiveRow([
                        ft.Column([
                            ft.Card(
                                ft.Container(
                                    ft.Column([
                                        ft.ResponsiveRow([
                                            self.tabela_produtos
                                        ])
                                    ], scroll=ft.ScrollMode.ALWAYS),
                                    padding=ft.padding.only(top=10, bottom=10), expand=True
                                ),expand=True,
                                elevation=5
                            )
                        ], col=6),
                        ft.Column([
                            ft.Card(
                                ft.Container(
                                    ft.Column([
                                        ft.ResponsiveRow([
                                            self.tabela_cotacao
                                        ])
                                    ], scroll=ft.ScrollMode.ALWAYS),
                                    padding=ft.padding.only(top=10, bottom=10), expand=True
                                ),
                                expand=True, elevation=5
                            )
                        ], col=6)
                    ]), expand=True
                )
            ]), width=650
        )
        self.actions = [
            ft.TextButton(
                text="Avançar",
                on_click=self.avancar,
                style=ft.ButtonStyle(bgcolor={ft.ControlState.HOVERED: ft.Colors.GREEN_100})
            ),
            ft.TextButton(
                text="Cancelar",
                on_click=lambda e: self.page.close(self),
                style=ft.ButtonStyle(bgcolor={ft.ControlState.HOVERED: ft.Colors.RED_100})
            )
        ]

    async def avancar(self, e: ft.ControlEvent) -> None:
        if self.lista_cotacao:
            self.segunda_tabela_produtos = ft.DataTable(
                columns=[
                    ft.DataColumn(ft.Text("Nome"))
                ],
                rows=[
                    ft.DataRow(
                        cells=[ft.DataCell(ft.Text(linha.cells[1].content.value))]
                    )
                    for linha in self.tabela_cotacao.rows
                ],
                col=12
            )

            self.tabela_fornecedor = self.tabela_fornecedor = ft.DataTable(
                columns=[
                    ft.DataColumn(ft.Text("Nome")),
                    ft.DataColumn(ft.Text("Apagar"))
                ]
            )

            self.content = ft.Container(
                ft.Stack([
                    ft.Container(
                        ft.ResponsiveRow([
                            ft.Column([
                                ft.Card(
                                    ft.Container(
                                        ft.Column([
                                            ft.ResponsiveRow([
                                                self.segunda_tabela_produtos
                                            ])
                                        ], scroll=ft.ScrollMode.ALWAYS),
                                        padding=ft.padding.only(top=10, bottom=10), expand=True
                                    ),expand=True, elevation=5
                                )
                            ], col=6),
                            ft.Column([
                                ft.Card(
                                    ft.Container(
                                        ft.Column([
                                            ft.ResponsiveRow([
                                                self.tabela_fornecedor
                                            ])
                                        ], scroll=ft.ScrollMode.ALWAYS),
                                        padding=ft.padding.only(top=10, bottom=10), expand=True
                                    ),
                                    expand=True, elevation=5
                                )
                            ], col=6)
                        ]), expand=True
                    ),
                    self.dialogo
                ], alignment=ft.alignment.center), width=650
            )
            self.actions.pop(0)
            self.actions.insert(
                0,
                ft.TextButton(
                    text="Criar",
                    on_click=self.criar_planilha,
                    style=ft.ButtonStyle(bgcolor={ft.ControlState.HOVERED: ft.Colors.GREEN_100})
                )
            )
            self.update()
            await self.buscar_fornecedores()

    async def criar_planilha(self, e: ft.ControlEvent) -> None:
        try:
            planilha = PlanilhaCotacao(self.segunda_tabela_produtos.rows, self.tabela_fornecedor.rows)
            planilha.criar()
        except Exception as e:
            self.dialogo.generico(ft.Icons.ERROR, "Houve um erro ao criar")
        else:
            self.dialogo.generico(ft.Icons.CHECK, "Planilha criada")
        finally:
            await asyncio.sleep(2)
            self.dialogo.limpar()

    async def buscar_fornecedores(self) -> None:
        lista_fonecedores = set()
        for id in self.lista_cotacao:
            controle = ControleItem(ModeloItem(id=id), None)
            fornecedores = await controle.buscar_fornecedores_relacao()
            if fornecedores:
                for fornecedor in fornecedores:
                    lista_fonecedores.add(fornecedor[2])
        self.adicionar_fornecedores(lista_fonecedores)

    def adicionar_fornecedores(self, fornecedores: list) -> None:
        self.tabela_fornecedor.rows = [
            ft.DataRow(
                cells=[
                    ft.DataCell(ft.Text(nome)),
                    ft.DataCell(
                        ft.IconButton(
                            ft.Icons.DELETE,
                            on_click=lambda e, nome=nome: self.apagar_fornecedor(nome)
                        )
                    )
                ]
            )
            for nome in fornecedores
        ]
        self.tabela_fornecedor.update()

    def apagar_fornecedor(self, nome: str) -> None:
        for linha in self.tabela_fornecedor.rows:
            if linha.cells[0].content.value == nome:
                self.tabela_fornecedor.rows.remove(linha)
        self.tabela_fornecedor.update()

    def adicionar_linha_tabela_itens(self, indice: int) -> None:
        dados = self.dados[indice]
        self.tabela_produtos.rows.append(
            ft.DataRow(
                cells=[
                    ft.DataCell(ft.Text(dados[0]), visible=False),
                    ft.DataCell(ft.Text(dados[1])),
                    ft.DataCell(
                        ft.IconButton(
                            ft.Icons.ADD,
                            on_click=lambda e, id=dados[0]: self.adicionar_item_tabela_cotacao(id)
                        )
                    )
                ]
            )
        )
        self.tabela_produtos.update()

    def adicionar_item_tabela_cotacao(self, id: int) -> None:
        if id not in self.lista_cotacao:
            self.lista_cotacao.append(id)
            for dado in self.dados:
                if dado[0] == id:
                    self.tabela_cotacao.rows.append(
                        ft.DataRow(
                            cells=[
                                ft.DataCell(ft.Text(dado[0]), visible=False),
                                ft.DataCell(ft.Text(dado[1])),
                                ft.DataCell(
                                    ft.IconButton(
                                        ft.Icons.DELETE,
                                        on_click=lambda e, id=dado[0]: self.apagar_item_tabela_cotacao(id)
                                    )
                                )
                            ]
                        )
                    )
                    self.tabela_cotacao.update()

    def apagar_registros_tabela_produtos(self) -> None:
        self.tabela_produtos.rows.clear()
        self.tabela_produtos.update()

    def apagar_item_tabela_cotacao(self, id: int) -> None:
        self.lista_cotacao.remove(id)
        for linha in self.tabela_cotacao.rows:
            if linha.cells[0].content.value == id:
                self.tabela_cotacao.rows.remove(linha)
        self.tabela_cotacao.update()

    def filtrar_categoria(self, e: ft.ControlEvent) -> None:
        self.adicionar_registro_tabela_itens(e.data)

    def adicionar_registro_tabela_itens(self, categoria: str) -> None:
        self.apagar_registros_tabela_produtos()
        for i, dado in enumerate(self.dados):
            if categoria == dado[2]:
                self.adicionar_linha_tabela_itens(i)

    def adicionar_categorias_dropdown(self) -> None:
        if self.dados:
            cats = list(set(dado[2] for dado in self.dados))
            self.dropdown_categoria.options = [
                ft.dropdown.Option(value)
                for value in cats
            ]
            self.dropdown_categoria.value = cats[0]
            self.dropdown_categoria.update()
            self.adicionar_registro_tabela_itens(cats[0])

    async def ler_dados(self) -> None:
        bd = BancoDeDados("db_app6.db")
        self.dados = await bd.fetch_all("SELECT id, nome, categoria FROM produto;")
        self.adicionar_categorias_dropdown()

    def did_mount(self) -> None:
        self.page.run_task(self.ler_dados)


class PaginaPrincipal(ft.Container):
    def __init__(self) -> None:
        super().__init__(expand=True)

    def atualizar_conteudo(self, conteudo: ft.Control, atualizar=True) -> None:
        self.content = conteudo
        if atualizar:
            self.update()


class Aplicativo(ft.Container):
    def __init__(self) -> None:
        super().__init__(expand=True)
        self.criar_botoes()
        self.pagina_principal = PaginaPrincipal()
        self.barra_menu_voltar = ft.AppBar(
            bgcolor=ft.Colors.PRIMARY_CONTAINER,
            leading=ft.IconButton(icon=ft.Icons.ARROW_BACK, on_click=self.voltar_pagina_itens)
        )
        self.controle_pagina = ControlePagina(self.pagina_principal, self.barra_menu_voltar)
        self.content = self.pagina_principal
        self.pagina_itens = PaginaItens(self.controle_pagina)
        self.controle_grade_item = ControleGradeItem(self.pagina_itens)

    def criar_botoes(self) -> None:
        self.botao_add_item = ft.PopupMenuItem(
            text="Add Produto",
            icon=ft.Icons.ADD_CIRCLE_OUTLINE,
            on_click=self.abrir_janela_add_item
        )
        self.botao_add_fornecedor = ft.PopupMenuItem(
            text="Add Fornecedor",
            icon=ft.Icons.PERSON_ADD_ROUNDED,
            on_click=self.abrir_janela_add_fornecedor
        )

    def criar_barra_menu(self) -> None:
        self.barra_pesquisa = BarraPesquisa(self.controle_grade_item)
        self.filtro_categoria = ft.PopupMenuButton(
            icon=ft.Icons.FILTER_ALT,
            items=[
                ft.PopupMenuItem(text=cat, on_click=self.filtrar_categoria, checked=not i)
                for i, cat in enumerate(["Todos", "Carnes", "Condimentos", "Frios", "Verdura", "Legume", "Fruta", "Limpeza"])
            ],
            icon_color=ft.Colors.BLACK87
        )
        self.barra_menu_principal = ft.AppBar(
            center_title=False,
            bgcolor=ft.Colors.PRIMARY_CONTAINER,
            actions=[
                self.barra_pesquisa,
                ft.VerticalDivider(opacity=0),
                self.filtro_categoria,
                ft.VerticalDivider(),
                ft.IconButton(
                    icon=ft.Icons.BALLOT,
                    icon_color=ft.Colors.BLACK87,
                    tooltip="Criar Lista de Compras",
                    on_click=self.pagina_lista_compras
                ),
                ft.IconButton(
                    icon=ft.Icons.DASHBOARD_ROUNDED,
                    icon_color=ft.Colors.BLACK87,
                    tooltip="Dashboard",
                    on_click=self.pagina_dashboad
                ),
                ft.IconButton(
                    icon=ft.Icons.PERSON,
                    icon_color=ft.Colors.BLACK87,
                    tooltip="Fornecedores",
                    on_click=self.pagina_fornecedores
                ),
                ft.VerticalDivider(),
                ft.PopupMenuButton(
                    items=[
                        self.botao_add_item,
                        self.botao_add_fornecedor,
                        ft.PopupMenuItem(),
                        ft.PopupMenuItem(
                            text="Cotação",
                            icon=ft.Icons.ATTACH_MONEY_ROUNDED,
                            on_click=self.abrir_janela_cotacao
                        )
                    ],
                    icon_color=ft.Colors.BLACK87
                )
            ],
        )

        self.page.appbar = self.barra_menu_principal

    def filtrar_categoria(self, e: ft.ControlEvent) -> None:
        for item in self.filtro_categoria.items:
            item.checked = item.text == e.control.text
            item.update()
        self.controle_grade_item.fitrar_categoria(e.control.text)

    def voltar_pagina_itens(self, e: ft.ControlEvent) -> None:
        self.page.overlay.clear()
        self.barra_menu_voltar.actions.clear()
        self.barra_menu_voltar.title = None
        self.barra_menu_voltar.update()
        self.page.appbar = self.barra_menu_principal
        self.pagina_principal.atualizar_conteudo(self.pagina_itens, atualizar=False)
        self.page.update()

    def pagina_dashboad(self, e: ft.ControlEvent) -> None:
        pagina = PaginaDashboard()
        self.controle_pagina.alterar_para_barra_voltar()
        self.controle_pagina.add_acao_barra(pagina.botoes_calendario())
        self.controle_pagina.atualizar_pagina(pagina)

    def pagina_fornecedores(self, e: ft.ControlEvent) -> None:
        pagina = PaginaFornecedores()
        self.controle_pagina.alterar_para_barra_voltar()
        self.controle_pagina.atualizar_pagina(pagina)

    def pagina_lista_compras(self, e: ft.ControlEvent) -> None:
        janela = JanelaListaCompras(self.controle_pagina)
        self.page.open(janela)

    def abrir_janela_add_item(self, e: ft.ControlEvent) -> None:
        janela = JanelaAdcionarItem(self.controle_grade_item)
        self.page.open(janela)

    def abrir_janela_add_fornecedor(self, e: ft.ControlEvent) -> None:
        janela = JanelaAdcionarFornecedor()
        self.page.open(janela)

    def abrir_janela_cotacao(self, e: ft.ControlEvent) -> None:
        janela = JanelaCotacao()
        self.page.open(janela)

    def did_mount(self) -> None:
        self.criar_barra_menu()
        self.page.run_task(self.pagina_itens.criar_cards_itens)
        self.controle_pagina.atualizar_pagina(self.pagina_itens)


def main(page: ft.Page) -> None:
    os.chdir("/home/luiz/gestor_compras/src")
    # Configura a página
    # page.window.title_bar_hidden = True
    # page.window.frameless = True
    # page.window.full_screen = True
    page.window.min_height = 700
    page.window.min_width = 1100
    page.locale_configuration = ft.LocaleConfiguration(
        supported_locales=[ft.Locale("en", "US"), ft.Locale("pt", "BR")],
        current_locale=ft.Locale("pt", "BR")
    )

    aplicativo = Aplicativo()
    page.add(aplicativo)
    page.update()


if __name__ == "__main__":
    ft.app(target=main, assets_dir="assets")