import flet as ft
import locale
import openpyxl
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from datetime import datetime

from acessorios import Utilidades, JanelaNotificacao
from modelos import ModeloItem
from controles import ControleItem

locale.setlocale(locale.LC_ALL, "pt_BR.UTF-8")


class ControleTabelas:
    def __init__(self, tabela_produtos, tabela_fornecedor, painel_consumo):
        self.tabela_fornecedor = tabela_fornecedor
        self.tabela_produtos = tabela_produtos
        self.painel_consumo = painel_consumo

    async def ler_fornecedores(self, id_produto: int) -> None:
        await self.tabela_fornecedor.ler_fornecedores(id_produto)

    def adicionar_fornecedor(self, id: int, nome: str, marca: str, preco: float) -> None:
        self.tabela_produtos.adicionar_fornecedor(id, nome, marca, preco)

    async def ler_consumo_e_infos(self, id_produto: int) -> None:
        await self.painel_consumo.ler_consumo_e_infos(id_produto)

    def adicionar_quantidade(self, id: int, quantidade: str, medida: str) -> None:
        self.tabela_produtos.adicionar_quantidade(id, quantidade, medida)

    def atualizar_preco(self, id: int, preco: float) -> None:
        self.tabela_produtos.atualizar_preco(id, preco)


class JanelaEdicao(ft.AlertDialog):
    def __init__(self, controle: ControleTabelas, id_produto: int) -> None:
        super().__init__(modal=True)
        self.controle = controle
        self.id_produto = id_produto
        self.field_valor = ft.TextField(label="Valor", width=200, prefix_text="R$ ")
        self.content = ft.Row([
            self.field_valor
        ])
        self.actions = [
            ft.TextButton(text="Atualizar", on_click=self.atualizar),
            ft.TextButton(text="Cancelar", on_click=lambda e: self.page.close(self))
        ]

    async def atualizar(self, e: ft.ControlEvent) -> None:
        preco = self.field_valor.value.replace(",", ".")
        self.controle.atualizar_preco(self.id_produto, float(preco))
        self.page.close(self)


class TabelaProdutos(ft.DataTable):
    def __init__(self) -> None:
        super().__init__(
            col=12,
            columns=[
                ft.DataColumn(ft.Text("id"), visible=False),
                ft.DataColumn(ft.Text("Nome")),
                ft.DataColumn(ft.Text("Fornecedor")),
                ft.DataColumn(ft.Text("Marca")),
                ft.DataColumn(ft.Text("Preço")),
                ft.DataColumn(ft.Text("Quantidade")),
                ft.DataColumn(ft.Text("Valor Operação"))
            ],
            column_spacing=10,
            heading_row_height=40
        )
        self.controle = None

    def adicionar_linha(self, dados: list) -> None:
        async def editar_fornecedor_ao_clicar(e: ft.ControlEvent, id: int=dados[0]) -> None:
            await self.editar_fornecedor(id)

        async def editar_consumo_e_infos_ao_clicar(e: ft.ControlEvent, id: int=dados[0]) -> None:
            await self.editar_consumo_e_infos(id)

        def editar_preco_ao_clicar(e: ft.ControlEvent, id: int=dados[0]) -> None:
            self.editar_preco(id)

        self.rows.append(
            ft.DataRow(
                cells=[
                    ft.DataCell(ft.Text(dados[0]), visible=False),
                    ft.DataCell(ft.Text(dados[1])),
                    ft.DataCell(ft.Text(dados[2]), show_edit_icon=True, on_tap=editar_fornecedor_ao_clicar),
                    ft.DataCell(ft.Text(dados[3])),
                    ft.DataCell(ft.Text(dados[4]), show_edit_icon=True, on_tap=editar_preco_ao_clicar),
                    ft.DataCell(ft.Text(dados[5]), show_edit_icon=True, on_tap=editar_consumo_e_infos_ao_clicar),
                    ft.DataCell(ft.Text(dados[6]))
                ]
            )
        )

    async def editar_fornecedor(self, id: int) -> None:
        await self.controle.ler_fornecedores(id)

    async def editar_consumo_e_infos(self, id: int) -> None:
        await self.controle.ler_consumo_e_infos(id)

    def editar_preco(self, id: int) -> None:
        janela = JanelaEdicao(self.controle, id)
        self.page.open(janela)

    def adicionar_fornecedor(self, id: int, nome: str, marca: str, preco: float) -> None:
        for row in self.rows:
            if int(id) == int(row.cells[0].content.value):
                row.cells[2].content.value = nome[:30]
                row.cells[3].content.value = marca[:20]
                row.cells[4].content.value = locale.currency(preco, grouping=True)
                row.cells[2].content.update()
                row.cells[3].content.update()
                row.cells[4].content.update()

    def adicionar_quantidade(self, id: int, quantidade: str, medida: str) -> None:
        for row in self.rows:
            if int(id) == int(row.cells[0].content.value) and quantidade:
                valor_opercao = self.calcular_valor_operacao(row, quantidade)
                row.cells[5].content.value = self.formatar_quantidade(quantidade, medida)
                row.cells[6].content.value = valor_opercao
                row.cells[5].content.update()
                row.cells[6].content.update()

    def atualizar_preco(self, id: int, preco: float) -> None:
        for row in self.rows:
            if int(id) == int(row.cells[0].content.value):
                row.cells[4].content.value = locale.currency(preco, grouping=True)
                row.cells[4].content.update()

    def calcular_valor_operacao(self, linha: ft.DataRow, quantidade: str) -> str:
        preco = linha.cells[4].content.value
        preco = float(preco.replace("R$", "").strip().replace(".", "").replace(",", "."))
        valor_operacao = float(quantidade) * preco
        return locale.currency(valor_operacao, grouping=True)

    def formatar_quantidade(self, quantidade: str, medida: str) -> str:
        return f"{str(quantidade).replace(".", ",")} {Utilidades.encurtar_medida(medida)}"

    def definir_controle(self, controle: ControleTabelas) -> None:
        self.controle = controle


class TabelaFornecedor(ft.DataTable):
    def __init__(self) -> None:
        super().__init__(
            col=12,
            columns=[
                ft.DataColumn(ft.Text("Nome")),
                ft.DataColumn(ft.Text("Marca")),
                ft.DataColumn(ft.Text("Preço")),
                ft.DataColumn(ft.Text("Adicionar"))
            ],
            column_spacing=10,
            heading_row_height=40
        )
        self.controle = None
        self.controle_produto = None

    def adicionar_linha(self, dado: list) -> None:
        def adicionar_ao_clicar(
            e: ft.ControlEvent,
            nome: str=dado[2],
            marca: str=dado[4],
            preco: float=dado[3]
        ) -> None:
            self.adicionar_fornecedor(nome, marca, preco)

        self.rows.append(
            ft.DataRow(
                cells=[
                    ft.DataCell(ft.Text(dado[2][:30])),
                    ft.DataCell(ft.Text(dado[4][:20])),
                    ft.DataCell(ft.Text(locale.currency(dado[3], grouping=True))),
                    ft.DataCell(
                        ft.IconButton(
                            ft.Icons.ADD,
                            on_click=adicionar_ao_clicar
                        )
                    )
                ]
            )
        )

    def adicionar_registros(self, dados: list) -> None:
        self.rows.clear()
        if dados:
            for dado in dados:
                self.adicionar_linha(dado)
            self.update()

    async def ler_fornecedores(self, id_produto: int) -> None:
        self.controle_produto = ControleItem(ModeloItem(id=id_produto), None)
        resultado = await self.controle_produto.buscar_fornecedores_relacao()
        self.adicionar_registros(self.ordenar_resultado(resultado))

    def ordenar_resultado(self, dados: list) -> None:
        return sorted(dados, key=lambda x: x[3])

    def adicionar_fornecedor(self, nome: str, marca: str, preco: str) -> None:
        self.controle.adicionar_fornecedor(self.controle_produto.id, nome, marca, preco)

    def definir_controle(self, controle: ControleTabelas) -> None:
        self.controle = controle


class TabelaConsumo(ft.DataTable):
    def __init__(self) -> None:
        super().__init__(
            columns=[
                ft.DataColumn(ft.Text(value="Dia")),
                ft.DataColumn(ft.Text(value="Consumo"))
            ],
            rows=[
                ft.DataRow(
                    cells=[
                        ft.DataCell(ft.Text(value=dia)),
                        ft.DataCell(
                            ft.Text(value="0"),
                            data=i
                        )
                    ]
                )
                for i, dia in enumerate(["Segunda-Feira", "Terça-Feira", "Quarta-Feira", "Quinta-Feira", "Sexta-Feira", "Sábado", "Domingo"], start=1)
            ],
            col=12,
            heading_row_height=40
        )

    def adicionar_registro(self, dados: list, medida: str) -> None:
        medida = Utilidades.encurtar_medida(medida)
        if dados:
            for dado in dados:
                for row in self.rows:
                    if int(row.cells[1].data) == int(dado[0]):
                        row.cells[1].content.value = f"{self.formatar_quantidade(dado[1])} {medida}"
                        row.update()

    async def ler_consumo(self, id_produto: int, medida: str) -> None:
        controle_produto = ControleItem(ModeloItem(id=id_produto), None)
        resultado = await controle_produto.obter_dados_consumo()
        self.adicionar_registro(resultado, medida)

    def formatar_quantidade(self, quantidade: str) -> str:
        return quantidade.replace(".", ",")


class Acondicionamento(ft.Container):
    def __init__(self) -> None:
        super().__init__(
            padding=ft.padding.all(10)
        )
        self.field_armazenamento = ft.TextField(label="Capacidade", value=0, width=205)
        self.field_dias = ft.TextField(label="Dias Máximo", value=0, width=200, suffix_text="Dias")
        self.content = ft.Column([
            ft.Row([
                self.field_armazenamento,
                self.field_dias
            ])
        ])

    def atualizar_valores(self, dados: list, medida: str) -> None:
        self.definir_sufixo_field(medida)
        self.field_armazenamento.value = self.formatar_quantidade(dados[0]) if dados else 0
        self.field_dias.value = dados[1] if dados else 0
        self.field_armazenamento.update()
        self.field_dias.update()

    def definir_sufixo_field(self, medida: str) -> None:
        self.field_armazenamento.suffix_text = Utilidades.encurtar_medida(medida)
        self.field_armazenamento.update()

    async def ler_dados(self, id_produto: int, medida: str) -> None:
        controle_produto = ControleItem(ModeloItem(id=id_produto), None)
        resultado = await controle_produto.obter_dados_infos()
        self.atualizar_valores(resultado, medida)

    def formatar_quantidade(self, quantidade: str) -> str:
        return quantidade.replace(".", ",")


class PainelInfos(ft.Container):
    def __init__(self) -> None:
        super().__init__()
        self.tabela_consumo = TabelaConsumo()
        self.acondicionamento = Acondicionamento()
        self.id_produto = None
        self.controle = None
        self.criar_conteudo()

    def criar_conteudo(self) -> None:
        self.field_qtd = ft.TextField(label="Quantidade", width=200, border="underline")
        self.content = ft.Column([
            ft.Container(
                ft.Tabs(
                    selected_index=0,
                    animation_duration=300,
                    tabs=[
                        ft.Tab(
                            text="Consumo",
                            content=ft.Container(
                                ft.Column([
                                    ft.ResponsiveRow([
                                        self.tabela_consumo
                                    ])
                                ], height=235, scroll=ft.ScrollMode.ALWAYS)
                            )
                        ),
                        ft.Tab(
                            text="Acondicionamento",
                            content=ft.Container(
                                self.acondicionamento
                            )
                        )
                    ]
                ),
                expand=True
            ),
            ft.Divider(height=6),
            ft.Container(
                ft.Row([
                    self.field_qtd,
                    ft.IconButton(ft.Icons.ADD, on_click=self.definir_quantidade)
                ], alignment=ft.MainAxisAlignment.END),
                padding=ft.padding.only(right=5, bottom=7)
            )
        ])

    async def ler_consumo_e_infos(self, id_produto: int) -> None:
        self.id_produto = id_produto
        self.medida = await self.obter_medida_produto(id_produto)
        await self.tabela_consumo.ler_consumo(id_produto, self.medida)
        await self.acondicionamento.ler_dados(id_produto, self.medida)

    async def obter_medida_produto(self, id_produto: int) -> None:
        controle_produto = ControleItem(ModeloItem(id=id_produto), None)
        medida = await controle_produto.obter_medida()
        return medida[0]

    def definir_quantidade(self, e: ft.ControlEvent) -> None:
        quantidade = self.field_qtd.value.replace(",", ".")
        self.controle.adicionar_quantidade(self.id_produto, quantidade, self.medida)
        self.limpar_campo()

    def limpar_campo(self) -> None:
        self.field_qtd.value = None
        self.field_qtd.update()

    def definir_controle(self, controle: ControleTabelas) -> None:
        self.controle = controle


class PlanilhaListaCompras:
    def __init__(self, linhas_tabela: ft.DataRow, nome_arquivo: str="compras.xlsx") -> None:
        self.linhas_tabela = linhas_tabela
        self.nome_arquivo = nome_arquivo
        self.colunas = ["Nome", "Fornecedor", "Marca", "Preço", "Quantidade", "Valor Op."]
        self.dims = [20, 20, 20, 15, 15, 15]
        self.book = openpyxl.Workbook()

    def criar(self) -> None:
        lista_produtos = self.extrair_valores()
        ws = self.book.active
        ws.append(self.colunas)
        for produto in lista_produtos:
            ws.append(produto)

        valor_total = self.valor_total(lista_produtos)
        ws.append(["valor total", "-", "-", "-", "-", valor_total])

        dim_holder = DimensionHolder(worksheet=ws)

        for col in range(ws.min_column, ws.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=self.dims[col-1])

        ws.column_dimensions = dim_holder
        
        self.adicionar_data_ao_nome()
        self.book.save(self.nome_arquivo)

    def extrair_valores(self) -> list:
        return [
            [linha.cells[i].content.value for i in range(1, 7)]
            for linha in self.linhas_tabela
        ]
    
    def valor_total(self, lista_produtos: list) -> str:
        valor_total = 0
        for produto in lista_produtos:
            valor_total += self.formatar_preco(produto[5])
        return locale.currency(valor_total, grouping=True)

    def formatar_preco(self, preco: str) -> float:
        return float(preco.replace("R$", "").strip().replace(",", "."))
    
    def adicionar_data_ao_nome(self):
        nome_list = self.nome_arquivo.split(".")
        nome_full = nome_list[0] + f"_{datetime.now().strftime("%d_%m_%Y")}." + nome_list[1]
        self.nome_arquivo = nome_full


class AgentePreenchedor:
    def __init__(self, infos_produtos: list, controle_tabelas: ControleTabelas):
        self.infos_produtos = infos_produtos
        self.controle_tabelas = controle_tabelas

    async def preencher(self):
        for produto in self.infos_produtos:
            controle = ControleItem(ModeloItem(produto[0]), None)
            fornecedor = await self.buscar_melhor_fornecedor(controle)
            infos = await self.obter_infos(controle)
            self.adicionar_fornecedor(fornecedor)
            self.adicionar_quantidade(infos)

    def adicionar_fornecedor(self, fornecedor: list):
        self.controle_tabelas.adicionar_fornecedor(fornecedor[0], fornecedor[1], fornecedor[3], fornecedor[2])

    def adicionar_quantidade(self, infos):
        if all(infos):
            qtd = self.calcular_quantidade(infos)
            qtd = self.formatar_quantidade(qtd, infos[2])
        else:
            qtd = 0
        self.controle_tabelas.adicionar_quantidade(infos[0], qtd, infos[2])

    async def buscar_melhor_fornecedor(self, controle: ControleItem):
        fornecedores = await controle.buscar_fornecedores_relacao()
        if fornecedores:
            fornecedor = sorted(fornecedores, key=lambda x: x[3])[0]
            return (controle.id, fornecedor[2], fornecedor[3], fornecedor[4])
        else:
            return (controle.id, "-", 0, "-")
        
    async def obter_infos(self, controle: ControleItem):
        infos = await controle.obter_dados_infos()
        medida = await controle.obter_medida()
        return (controle.id, infos[0], medida[0], infos[2])
    
    def calcular_quantidade(self, infos):
        if all(infos):
            qtd = float(infos[1])
            qtd_media = infos[3]
            calc = qtd - (qtd * (qtd_media/100))
            return round(calc, 3)
            return 0
        
    def formatar_quantidade(self, qtd, medida):
        if medida == "unidade":
            qtd = int(qtd)
        else:
            qtd_split = str(qtd).split(".")
            zeros = "0" * (3 - len(qtd_split[1]))
            qtd = qtd_split[0] + "." + qtd_split[1] + zeros
        return qtd


class PaginaListacompras(ft.Container):
    def __init__(self, infos_produtos: list, preencher_automatico: bool) -> None:
        super().__init__(expand=True)
        self.infos_produtos = infos_produtos
        self.preencher_automatico = preencher_automatico
        self.tabela_produtos = TabelaProdutos()
        self.tabela_fornecedores = TabelaFornecedor()
        self.painel_infos = PainelInfos()
        self.controle_tabelas = ControleTabelas(
            self.tabela_produtos, self.tabela_fornecedores, self.painel_infos
        )
        self.agente_preenchedor = AgentePreenchedor(infos_produtos, self.controle_tabelas)
        self.content = ft.ResponsiveRow([
            ft.Column([
                ft.Card(
                    ft.Container(
                        ft.Column([
                            ft.ResponsiveRow([
                                self.tabela_produtos
                            ])
                        ], scroll=ft.ScrollMode.ALWAYS),
                        padding=ft.padding.only(top=10, bottom=10), expand=True
                    ),
                    expand=True, elevation=5
                )
            ], col=8),
            ft.Column([
                ft.Card(
                    ft.Container(
                        ft.Column([
                            ft.ResponsiveRow([
                                self.tabela_fornecedores
                            ])   
                        ], scroll=ft.ScrollMode.ALWAYS),
                        padding=ft.padding.only(top=10, bottom=10), expand=True
                    ), expand=True, elevation=5
                ),
                ft.Card(
                    self.painel_infos, expand=True, elevation=5
                )
            ], col=4)
        ])

    def adicionar_registros_primarios(self) -> None:
        for id, nome in self.infos_produtos:
            self.tabela_produtos.adicionar_linha([id, nome, "-", "-", "R$ 00,00", "0", "R$ 00,00"])
        self.tabela_produtos.update()

    def criar_planilha(self, e: ft.ControlEvent) -> None:
        planilha = PlanilhaListaCompras(self.tabela_produtos.rows)
        try:
            planilha.criar()
        except Exception as e:
            janela = JanelaNotificacao("Houve um erro!", ft.icons.ERROR)
        else:
            janela = JanelaNotificacao("Tabela Criada", ft.icons.CHECK)
        self.page.open(janela)

    def did_mount(self) -> None:
        self.tabela_produtos.definir_controle(self.controle_tabelas)
        self.tabela_fornecedores.definir_controle(self.controle_tabelas)
        self.painel_infos.definir_controle(self.controle_tabelas)
        self.adicionar_registros_primarios()
        if self.preencher_automatico:
            self.page.run_task(self.agente_preenchedor.preencher)
